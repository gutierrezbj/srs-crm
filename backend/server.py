from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Response, Request, Depends, BackgroundTasks
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import csv
import io
import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI(title="System Rapid Solutions CRM")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============== MODELS ==============

# Lead stages
LEAD_STAGES = ["nuevo", "contactado", "calificado", "propuesta", "negociacion", "ganado", "perdido"]

# Dropdown options
SECTORES = [
    "IT / Soporte técnico",
    "Fotovoltaica / Energía",
    "Drones / Inspección",
    "Telecomunicaciones",
    "Retail",
    "Banca / Finanzas",
    "Industria / Manufactura",
    "Logística",
    "Otro"
]

SERVICIOS = [
    "Soporte IT Remoto",
    "Soporte IT Presencial",
    "Smart Hands / Remote Hands",
    "Managed Services (MSP)",
    "Inspección con Drones",
    "Termografía",
    "Topografía / Fotogrametría",
    "Cableado Estructurado",
    "Consultoría ENS",
    "Formación / Cursos",
    "Otro"
]

FUENTES = [
    "Apollo",
    "Web",
    "Referido",
    "Feria / Evento",
    "LinkedIn",
    "Licitación",
    "Otro"
]

URGENCIAS = [
    "Inmediata (< 1 mes)",
    "Corto plazo (1-3 meses)",
    "Medio plazo (3-6 meses)",
    "Largo plazo (6+ meses)",
    "Sin definir"
]

MOTIVOS_PERDIDA = [
    "Precio",
    "Timing",
    "Competencia",
    "No responde",
    "No califica",
    "Otro"
]

TIPOS_SEGUIMIENTO = [
    "Llamada",
    "Email",
    "Reunión",
    "Demo",
    "Propuesta",
    "Otro"
]

# Allowed users (initial, can be managed via admin panel)
ALLOWED_USERS = {
    "juancho@systemrapidsolutions.com": {"name": "JuanCho", "role": "admin"},
    "andros@systemrapidsolutions.com": {"name": "Andros", "role": "user"},
    "adriana@systemrapidsolutions.com": {"name": "Adriana", "role": "user"},
}

# Development mode - set to False for production
DEV_MODE = os.environ.get('DEV_MODE', 'true').lower() == 'true'

class UserBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str = "user"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: str

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    role: str = "user"

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None

class LeadBase(BaseModel):
    empresa: str
    contacto: str
    email: str  # Changed from EmailStr to allow empty/invalid emails in imported data
    telefono: Optional[str] = None
    cargo: Optional[str] = None
    sector: Optional[str] = None
    valor_estimado: float = 0.0
    etapa: str = "nuevo"
    notas: Optional[str] = None
    # New fields
    propietario: Optional[str] = None  # user_id del propietario
    servicios: Optional[List[str]] = []  # Multiple services
    fuente: Optional[str] = None
    urgencia: Optional[str] = "Sin definir"
    motivo_perdida: Optional[str] = None
    proximo_seguimiento: Optional[str] = None  # ISO date
    tipo_seguimiento: Optional[str] = None

class LeadCreate(LeadBase):
    pass

class LeadUpdate(BaseModel):
    empresa: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[str] = None  # Changed from EmailStr
    telefono: Optional[str] = None
    cargo: Optional[str] = None
    sector: Optional[str] = None
    valor_estimado: Optional[float] = None
    etapa: Optional[str] = None
    notas: Optional[str] = None
    propietario: Optional[str] = None
    servicios: Optional[List[str]] = None
    fuente: Optional[str] = None
    urgencia: Optional[str] = None
    motivo_perdida: Optional[str] = None
    proximo_seguimiento: Optional[str] = None
    tipo_seguimiento: Optional[str] = None

class Lead(LeadBase):
    model_config = ConfigDict(extra="ignore")
    lead_id: str
    fecha_creacion: datetime
    fecha_ultimo_contacto: Optional[datetime] = None
    dias_sin_actividad: int = 0
    created_by: str
    propietario_nombre: Optional[str] = None

class ActivityBase(BaseModel):
    tipo: str  # nota, llamada, email, reunion
    descripcion: str

class ActivityCreate(ActivityBase):
    pass

class Activity(ActivityBase):
    model_config = ConfigDict(extra="ignore")
    activity_id: str
    lead_id: str
    user_id: str
    user_name: str
    created_at: datetime

class EnrichRequest(BaseModel):
    email: str
    empresa: Optional[str] = None
    contacto: Optional[str] = None
    telefono: Optional[str] = None
    cargo: Optional[str] = None
    sector: Optional[str] = None

# ============== TIPOS SRS (pilares de servicio) ==============
TIPOS_SRS = [
    "IT / Soporte técnico",
    "Drones / Inspección", 
    "Telecomunicaciones",
    "Consultoría ENS",
    "Fotovoltaica / Energía",
    "Formación",
    "Otro"
]

# ============== OPORTUNIDADES PLACSP MODELS ==============
class OportunidadPLACSPBase(BaseModel):
    expediente: str
    adjudicatario: str
    nif: str
    importe: float
    objeto: str
    cpv: str
    score: int  # 0-100
    tipo_srs: str
    keywords: List[str] = []
    indicadores_dolor: List[str] = []
    fecha_adjudicacion: datetime
    fecha_fin_contrato: Optional[datetime] = None
    dias_restantes: Optional[int] = None
    url_licitacion: str
    url_pliego: Optional[str] = None
    organo_contratacion: str
    es_pyme: bool = False
    convertido_lead: bool = False
    fecha_deteccion: datetime
    # Estado de revisión del operador
    estado_revision: str = "nueva"  # nueva, revisada, descartada
    fecha_revision: Optional[datetime] = None
    revisado_por: Optional[str] = None
    # Pain Score y análisis
    pain_score: Optional[int] = None  # 0-100
    nivel_urgencia: Optional[str] = None  # critico, alto, medio, bajo
    pain_analysis: Optional[Dict[str, Any]] = None
    analisis_fecha: Optional[str] = None
    # Análisis exhaustivo del pliego
    analisis_pliego: Optional[Dict[str, Any]] = None
    analisis_pliego_fecha: Optional[str] = None
    tiene_it_confirmado: Optional[bool] = None
    # Datos enriquecidos del adjudicatario
    datos_adjudicatario: Optional[Dict[str, Any]] = None  # {nombre, nif, direccion, telefono, email, web, sector, empleados, fuente, fecha_enriquecimiento}

class OportunidadPLACSPCreate(OportunidadPLACSPBase):
    pass

class OportunidadPLACSP(OportunidadPLACSPBase):
    model_config = ConfigDict(extra="ignore")
    oportunidad_id: str
    ref_code: Optional[str] = None  # Código corto de referencia: "01", "02", etc.

class OportunidadSpotterImport(BaseModel):
    oportunidades: List[OportunidadPLACSPCreate]

# ============== AUTH HELPERS ==============

async def get_current_user(request: Request) -> UserResponse:
    """Get current user from session token cookie or Authorization header"""
    session_token = request.cookies.get("session_token")
    
    if not session_token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            session_token = auth_header.split(" ")[1]
    
    if not session_token:
        raise HTTPException(status_code=401, detail="No autenticado")
    
    session = await db.user_sessions.find_one(
        {"session_token": session_token},
        {"_id": 0}
    )
    
    if not session:
        raise HTTPException(status_code=401, detail="Sesión inválida")
    
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Sesión expirada")
    
    user = await db.users.find_one(
        {"user_id": session["user_id"]},
        {"_id": 0}
    )
    
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    
    return UserResponse(**user)

# ============== AUTH ROUTES ==============

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    """Exchange session_id from Emergent Auth for session data"""
    session_id = request.headers.get("X-Session-ID")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="Session ID requerido")
    
    # Exchange session_id with Emergent Auth
    async with httpx.AsyncClient() as client:
        auth_response = await client.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": session_id}
        )
        
        if auth_response.status_code != 200:
            raise HTTPException(status_code=401, detail="Sesión inválida de Emergent Auth")
        
        auth_data = auth_response.json()
    
    email = auth_data.get("email", "").lower()
    
    # In DEV_MODE, allow any email. In production, restrict to @systemrapidsolutions.com
    if not DEV_MODE:
        # Check domain restriction
        if not email.endswith("@systemrapidsolutions.com"):
            raise HTTPException(
                status_code=403, 
                detail="Solo cuentas @systemrapidsolutions.com permitidas"
            )
        
        # Check if user is in allowed list (case insensitive)
        allowed_user = ALLOWED_USERS.get(email)
        if not allowed_user:
            raise HTTPException(
                status_code=403,
                detail="Usuario no autorizado. Contacte al administrador."
            )
        user_role = allowed_user["role"]
        user_name = auth_data.get("name", allowed_user["name"])
    else:
        # DEV_MODE: Allow any Google account as admin for testing
        user_role = "admin"
        user_name = auth_data.get("name", email.split("@")[0])
    
    # Create or update user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    existing_user = await db.users.find_one({"email": email}, {"_id": 0})
    
    if existing_user:
        user_id = existing_user["user_id"]
        await db.users.update_one(
            {"email": email},
            {"$set": {
                "name": user_name,
                "picture": auth_data.get("picture"),
                "role": user_role
            }}
        )
    else:
        new_user = {
            "user_id": user_id,
            "email": email,
            "name": user_name,
            "picture": auth_data.get("picture"),
            "role": user_role,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(new_user)
    
    # Create session
    session_token = auth_data.get("session_token", f"session_{uuid.uuid4().hex}")
    expires_at = datetime.now(timezone.utc) + timedelta(days=7)
    
    await db.user_sessions.delete_many({"user_id": user_id})
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # Set cookie
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        max_age=7 * 24 * 60 * 60,
        path="/"
    )
    
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return UserResponse(**user)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: UserResponse = Depends(get_current_user)):
    """Get current authenticated user"""
    return current_user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    """Logout user and clear session"""
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_many({"session_token": session_token})
    
    response.delete_cookie(key="session_token", path="/")
    return {"message": "Sesión cerrada"}

# ============== LEADS ROUTES ==============

def calculate_days_without_activity(fecha_ultimo_contacto: Optional[datetime]) -> int:
    """Calculate days since last activity"""
    if not fecha_ultimo_contacto:
        return 999
    
    if isinstance(fecha_ultimo_contacto, str):
        fecha_ultimo_contacto = datetime.fromisoformat(fecha_ultimo_contacto)
    
    if fecha_ultimo_contacto.tzinfo is None:
        fecha_ultimo_contacto = fecha_ultimo_contacto.replace(tzinfo=timezone.utc)
    
    delta = datetime.now(timezone.utc) - fecha_ultimo_contacto
    return delta.days

@api_router.get("/leads", response_model=List[Lead])
async def get_leads(
    etapa: Optional[str] = None,
    sector: Optional[str] = None,
    propietario: Optional[str] = None,
    search: Optional[str] = None,
    seguimiento: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Get all leads with optional filters"""
    query = {}

    if etapa:
        query["etapa"] = etapa
    if sector:
        query["sector"] = sector
    if propietario:
        query["propietario"] = propietario
    if search:
        query["$or"] = [
            {"empresa": {"$regex": search, "$options": "i"}},
            {"contacto": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]

    # Filtro de seguimiento
    if seguimiento:
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today + timedelta(days=1)

        if seguimiento == "hoy":
            # Seguimientos programados para hoy
            query["proximo_seguimiento"] = {
                "$gte": today.isoformat(),
                "$lt": today_end.isoformat()
            }
        elif seguimiento == "atrasados":
            # Seguimientos vencidos (antes de hoy)
            query["proximo_seguimiento"] = {
                "$lt": today.isoformat(),
                "$ne": None
            }
        elif seguimiento == "proximos":
            # Próximos 7 días
            week_end = today + timedelta(days=7)
            query["proximo_seguimiento"] = {
                "$gte": today.isoformat(),
                "$lt": week_end.isoformat()
            }

    leads = await db.leads.find(query, {"_id": 0}).sort("fecha_creacion", -1).to_list(1000)
    
    # Get all users for propietario_nombre lookup
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1}).to_list(100)
    users_map = {u["user_id"]: u["name"] for u in users}
    
    for lead in leads:
        if isinstance(lead.get("fecha_creacion"), str):
            lead["fecha_creacion"] = datetime.fromisoformat(lead["fecha_creacion"])
        if lead.get("fecha_ultimo_contacto") and isinstance(lead["fecha_ultimo_contacto"], str):
            lead["fecha_ultimo_contacto"] = datetime.fromisoformat(lead["fecha_ultimo_contacto"])
        lead["dias_sin_actividad"] = calculate_days_without_activity(lead.get("fecha_ultimo_contacto"))
        # Add propietario name
        if lead.get("propietario"):
            lead["propietario_nombre"] = users_map.get(lead["propietario"], "")
        else:
            lead["propietario_nombre"] = None
        # Ensure servicios is a list
        if not lead.get("servicios"):
            lead["servicios"] = []
    
    return leads

@api_router.get("/leads/stats")
async def get_leads_stats(current_user: UserResponse = Depends(get_current_user)):
    """Get dashboard statistics"""
    leads = await db.leads.find({}, {"_id": 0}).to_list(1000)
    today = datetime.now(timezone.utc).date()
    
    # Count by stage
    stages_count = {stage: 0 for stage in LEAD_STAGES}
    total_pipeline = 0.0
    leads_without_activity = 0
    seguimientos_hoy = []
    
    for lead in leads:
        etapa = lead.get("etapa", "nuevo")
        stages_count[etapa] = stages_count.get(etapa, 0) + 1
        
        if etapa not in ["ganado", "perdido"]:
            total_pipeline += lead.get("valor_estimado", 0)
        
        dias = calculate_days_without_activity(lead.get("fecha_ultimo_contacto"))
        if dias > 7 and etapa not in ["ganado", "perdido"]:
            leads_without_activity += 1
        
        # Check for seguimientos today
        proximo = lead.get("proximo_seguimiento")
        if proximo:
            try:
                fecha_seg = datetime.fromisoformat(proximo.replace("Z", "+00:00")).date() if isinstance(proximo, str) else proximo.date()
                if fecha_seg == today:
                    seguimientos_hoy.append({
                        "lead_id": lead.get("lead_id"),
                        "empresa": lead.get("empresa"),
                        "contacto": lead.get("contacto"),
                        "tipo_seguimiento": lead.get("tipo_seguimiento", "Llamada"),
                        "proximo_seguimiento": proximo
                    })
            except:
                pass
    
    return {
        "stages_count": stages_count,
        "total_pipeline": total_pipeline,
        "leads_without_activity": leads_without_activity,
        "total_leads": len(leads),
        "seguimientos_hoy": seguimientos_hoy
    }

@api_router.get("/leads/export")
async def export_leads(
    etapa: Optional[str] = None,
    sector: Optional[str] = None,
    seguimiento: Optional[str] = None,
    format: str = "xlsx",
    current_user: UserResponse = Depends(get_current_user)
):
    """Export leads to Excel or CSV with optional filters"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Build query with filters
    query = {}
    if etapa:
        query["etapa"] = etapa
    if sector:
        query["sector"] = sector

    # Seguimiento filter
    if seguimiento:
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today + timedelta(days=1)
        if seguimiento == "hoy":
            query["proximo_seguimiento"] = {"$gte": today.isoformat(), "$lt": today_end.isoformat()}
        elif seguimiento == "atrasados":
            query["proximo_seguimiento"] = {"$lt": today.isoformat(), "$ne": None}
        elif seguimiento == "proximos":
            week_end = today + timedelta(days=7)
            query["proximo_seguimiento"] = {"$gte": today.isoformat(), "$lt": week_end.isoformat()}

    leads = await db.leads.find(query, {"_id": 0}).sort("fecha_creacion", -1).to_list(1000)

    # Get users for propietario lookup
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1}).to_list(100)
    users_map = {u["user_id"]: u["name"] for u in users}

    if format == "csv":
        # CSV export
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            "empresa", "contacto", "email", "telefono", "cargo",
            "sector", "valor_estimado", "etapa", "notas"
        ])
        writer.writeheader()
        for lead in leads:
            writer.writerow({
                "empresa": lead.get("empresa", ""),
                "contacto": lead.get("contacto", ""),
                "email": lead.get("email", ""),
                "telefono": lead.get("telefono", ""),
                "cargo": lead.get("cargo", ""),
                "sector": lead.get("sector", ""),
                "valor_estimado": lead.get("valor_estimado", 0),
                "etapa": lead.get("etapa", "nuevo"),
                "notas": lead.get("notas", "")
            })
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=leads_export.csv"}
        )

    # Excel export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Headers
    headers = [
        "Empresa", "Contacto", "Email", "Teléfono", "Cargo",
        "Sector", "Valor Estimado", "Etapa", "Propietario",
        "Urgencia", "Servicios", "Fuente", "Próximo Seguimiento", "Notas"
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Stage colors for conditional formatting
    stage_colors = {
        "nuevo": "64748B",
        "contactado": "3B82F6",
        "calificado": "06B6D4",
        "propuesta": "A855F7",
        "negociacion": "F59E0B",
        "ganado": "10B981",
        "perdido": "EF4444"
    }

    # Data rows
    for row_num, lead in enumerate(leads, 2):
        etapa_val = lead.get("etapa", "nuevo")
        propietario_id = lead.get("propietario", "")
        propietario_name = users_map.get(propietario_id, "") if propietario_id else ""
        servicios = ", ".join(lead.get("servicios", []))
        proximo_seg = lead.get("proximo_seguimiento", "")
        if proximo_seg:
            try:
                fecha = datetime.fromisoformat(proximo_seg.replace("Z", "+00:00"))
                proximo_seg = fecha.strftime("%Y-%m-%d")
            except:
                pass

        row_data = [
            lead.get("empresa", ""),
            lead.get("contacto", ""),
            lead.get("email", ""),
            lead.get("telefono", ""),
            lead.get("cargo", ""),
            lead.get("sector", ""),
            lead.get("valor_estimado", 0),
            etapa_val.capitalize(),
            propietario_name,
            lead.get("urgencia", ""),
            servicios,
            lead.get("fuente", ""),
            proximo_seg,
            lead.get("notas", "")
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            # Apply stage color to Etapa column
            if col == 8:
                color = stage_colors.get(etapa_val, "64748B")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = len(headers[col - 1])
        for row in range(2, len(leads) + 2):
            cell_value = str(ws.cell(row=row, column=col).value or "")
            if len(cell_value) > max_length:
                max_length = min(len(cell_value), 50)
        ws.column_dimensions[get_column_letter(col)].width = max_length + 2

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add filters
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(leads) + 1}"

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/leads/{lead_id}", response_model=Lead)
async def get_lead(lead_id: str, current_user: UserResponse = Depends(get_current_user)):
    """Get single lead by ID"""
    lead = await db.leads.find_one({"lead_id": lead_id}, {"_id": 0})
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    if isinstance(lead.get("fecha_creacion"), str):
        lead["fecha_creacion"] = datetime.fromisoformat(lead["fecha_creacion"])
    if lead.get("fecha_ultimo_contacto") and isinstance(lead["fecha_ultimo_contacto"], str):
        lead["fecha_ultimo_contacto"] = datetime.fromisoformat(lead["fecha_ultimo_contacto"])
    lead["dias_sin_actividad"] = calculate_days_without_activity(lead.get("fecha_ultimo_contacto"))
    
    # Get propietario name
    if lead.get("propietario"):
        owner = await db.users.find_one({"user_id": lead["propietario"]}, {"_id": 0, "name": 1})
        lead["propietario_nombre"] = owner["name"] if owner else None
    else:
        lead["propietario_nombre"] = None
    
    if not lead.get("servicios"):
        lead["servicios"] = []
    
    return Lead(**lead)

@api_router.post("/leads", response_model=Lead)
async def create_lead(lead: LeadCreate, current_user: UserResponse = Depends(get_current_user)):
    """Create new lead"""
    if lead.etapa not in LEAD_STAGES:
        raise HTTPException(status_code=400, detail=f"Etapa inválida. Usar: {LEAD_STAGES}")
    
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    lead_dict = lead.model_dump()
    # Ensure servicios is a list
    if not lead_dict.get("servicios"):
        lead_dict["servicios"] = []
    
    lead_doc = {
        "lead_id": lead_id,
        **lead_dict,
        "fecha_creacion": now.isoformat(),
        "fecha_ultimo_contacto": now.isoformat(),
        "created_by": current_user.user_id
    }
    
    await db.leads.insert_one(lead_doc)
    
    lead_doc["dias_sin_actividad"] = 0
    lead_doc["fecha_creacion"] = now
    lead_doc["fecha_ultimo_contacto"] = now
    
    # Get propietario name
    if lead_doc.get("propietario"):
        owner = await db.users.find_one({"user_id": lead_doc["propietario"]}, {"_id": 0, "name": 1})
        lead_doc["propietario_nombre"] = owner["name"] if owner else None
    else:
        lead_doc["propietario_nombre"] = None
    
    return Lead(**lead_doc)

@api_router.put("/leads/{lead_id}", response_model=Lead)
async def update_lead(
    lead_id: str, 
    lead_update: LeadUpdate, 
    current_user: UserResponse = Depends(get_current_user)
):
    """Update lead"""
    existing = await db.leads.find_one({"lead_id": lead_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    update_data = {}
    for k, v in lead_update.model_dump().items():
        if v is not None:
            update_data[k] = v
        elif k == "servicios" and lead_update.servicios is not None:
            update_data[k] = lead_update.servicios  # Allow empty list
    
    if "etapa" in update_data and update_data["etapa"] not in LEAD_STAGES:
        raise HTTPException(status_code=400, detail=f"Etapa inválida. Usar: {LEAD_STAGES}")
    
    if update_data:
        await db.leads.update_one({"lead_id": lead_id}, {"$set": update_data})
    
    updated = await db.leads.find_one({"lead_id": lead_id}, {"_id": 0})
    
    if isinstance(updated.get("fecha_creacion"), str):
        updated["fecha_creacion"] = datetime.fromisoformat(updated["fecha_creacion"])
    if updated.get("fecha_ultimo_contacto") and isinstance(updated["fecha_ultimo_contacto"], str):
        updated["fecha_ultimo_contacto"] = datetime.fromisoformat(updated["fecha_ultimo_contacto"])
    updated["dias_sin_actividad"] = calculate_days_without_activity(updated.get("fecha_ultimo_contacto"))
    
    # Get propietario name
    if updated.get("propietario"):
        owner = await db.users.find_one({"user_id": updated["propietario"]}, {"_id": 0, "name": 1})
        updated["propietario_nombre"] = owner["name"] if owner else None
    else:
        updated["propietario_nombre"] = None
    
    if not updated.get("servicios"):
        updated["servicios"] = []
    
    return Lead(**updated)

@api_router.patch("/leads/{lead_id}/stage")
async def update_lead_stage(
    lead_id: str,
    etapa: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """Update lead stage (for Kanban drag & drop)"""
    if etapa not in LEAD_STAGES:
        raise HTTPException(status_code=400, detail=f"Etapa inválida. Usar: {LEAD_STAGES}")
    
    result = await db.leads.update_one(
        {"lead_id": lead_id},
        {"$set": {"etapa": etapa}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    return {"message": "Etapa actualizada", "lead_id": lead_id, "etapa": etapa}

@api_router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str, current_user: UserResponse = Depends(get_current_user)):
    """Delete lead"""
    result = await db.leads.delete_one({"lead_id": lead_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    # Delete related activities
    await db.activities.delete_many({"lead_id": lead_id})
    
    return {"message": "Lead eliminado"}

# ============== ACTIVITIES ROUTES ==============

@api_router.get("/leads/{lead_id}/activities", response_model=List[Activity])
async def get_lead_activities(lead_id: str, current_user: UserResponse = Depends(get_current_user)):
    """Get activities for a lead"""
    activities = await db.activities.find(
        {"lead_id": lead_id}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    for act in activities:
        if isinstance(act.get("created_at"), str):
            act["created_at"] = datetime.fromisoformat(act["created_at"])
    
    return activities

@api_router.post("/leads/{lead_id}/activities", response_model=Activity)
async def create_activity(
    lead_id: str, 
    activity: ActivityCreate, 
    current_user: UserResponse = Depends(get_current_user)
):
    """Create activity for a lead"""
    lead = await db.leads.find_one({"lead_id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    now = datetime.now(timezone.utc)
    activity_id = f"act_{uuid.uuid4().hex[:12]}"
    
    activity_doc = {
        "activity_id": activity_id,
        "lead_id": lead_id,
        "user_id": current_user.user_id,
        "user_name": current_user.name,
        **activity.model_dump(),
        "created_at": now.isoformat()
    }
    
    await db.activities.insert_one(activity_doc)
    
    # Update lead's last contact date
    await db.leads.update_one(
        {"lead_id": lead_id},
        {"$set": {"fecha_ultimo_contacto": now.isoformat()}}
    )
    
    activity_doc["created_at"] = now
    return Activity(**activity_doc)

# ============== IMPORT/EXPORT ROUTES ==============

@api_router.post("/leads/parse")
async def parse_file(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    """Parse CSV/Excel file and return headers and preview rows"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")
    
    content = await file.read()
    headers = []
    rows = []
    
    try:
        if file.filename.endswith('.csv'):
            decoded = content.decode('utf-8')
            reader = csv.reader(io.StringIO(decoded))
            all_rows = list(reader)
            if all_rows:
                headers = all_rows[0]
                rows = all_rows[1:]
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers = [str(cell) if cell is not None else "" for cell in all_rows[0]]
                rows = [[str(cell) if cell is not None else "" for cell in row] for row in all_rows[1:]]
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o Excel.")
        
        # Return first 3 rows as preview
        preview = rows[:3] if len(rows) >= 3 else rows
        
        return {
            "headers": headers,
            "rows": rows,
            "preview": preview,
            "total_rows": len(rows)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")


class DuplicateCheckRequest(BaseModel):
    leads: List[dict]


@api_router.post("/leads/check-duplicates")
async def check_duplicates(
    data: DuplicateCheckRequest,
    current_user: UserResponse = Depends(get_current_user)
):
    """Check for duplicate leads by email or company name"""
    duplicates = []
    
    # Get all existing leads
    existing_leads = await db.leads.find({}, {"_id": 0}).to_list(10000)
    
    # Build indexes for fast lookup
    # Email index (only for non-empty emails)
    existing_by_email = {}
    for lead in existing_leads:
        email = str(lead.get("email", "")).lower().strip()
        # Normalize email: remove extra spaces, take first email if multiple
        email = email.split()[0] if email else ""  # Take first part if spaces
        email = email.split("/")[0].strip() if "/" in email else email  # Take first if "email / email"
        if email and "@" in email:
            existing_by_email[email] = lead
    
    # Company index (normalized)
    existing_by_company = {}
    for lead in existing_leads:
        empresa = str(lead.get("empresa", "")).lower().strip()
        if empresa:
            existing_by_company[empresa] = lead
    
    for idx, new_lead in enumerate(data.leads):
        # Normalize new lead's email
        raw_email = str(new_lead.get("email", "")).lower().strip()
        email = raw_email.split()[0] if raw_email else ""
        email = email.split("/")[0].strip() if "/" in email else email
        has_valid_email = email and "@" in email
        
        # Normalize new lead's company
        empresa = str(new_lead.get("empresa", "")).lower().strip()
        
        duplicate_found = False
        
        # Strategy 1: Match by email (if both have valid emails)
        if has_valid_email and email in existing_by_email:
            duplicates.append({
                "rowIndex": idx,
                "type": "exact",
                "newLead": new_lead,
                "existingLead": existing_by_email[email]
            })
            duplicate_found = True
        
        # Strategy 2: Match by company name (if email is empty/invalid OR same company different email)
        if not duplicate_found and empresa and empresa in existing_by_company:
            existing = existing_by_company[empresa]
            existing_email = str(existing.get("email", "")).lower().strip()
            
            # If new lead has no valid email, or existing has no valid email, or they're different
            # This catches: empty emails, same company different contact
            if not has_valid_email or not existing_email or "@" not in existing_email:
                duplicates.append({
                    "rowIndex": idx,
                    "type": "exact",  # Same company = exact duplicate
                    "newLead": new_lead,
                    "existingLead": existing
                })
            else:
                # Both have emails but different - possible duplicate (different contact same company)
                duplicates.append({
                    "rowIndex": idx,
                    "type": "possible",
                    "newLead": new_lead,
                    "existingLead": existing
                })
    
    return {"duplicates": duplicates}


class ImportMappedRequest(BaseModel):
    leads: List[dict]
    mapping: dict


@api_router.post("/leads/import-mapped")
async def import_mapped_leads(
    data: ImportMappedRequest,
    current_user: UserResponse = Depends(get_current_user)
):
    """Import leads with custom column mapping and duplicate handling"""
    imported = 0
    updated = 0
    skipped = 0
    errors = []
    
    for idx, lead_data in enumerate(data.leads):
        try:
            action = lead_data.pop("_action", "import")
            existing_id = lead_data.pop("_existingId", None)
            
            if action == "skip":
                skipped += 1
                continue
            
            # Clean and validate data
            lead_doc = {
                "empresa": str(lead_data.get("empresa", "")).strip() or "Sin empresa",
                "contacto": str(lead_data.get("contacto", "")).strip() or "Sin contacto",
                "email": str(lead_data.get("email", "")).strip().lower(),
                "telefono": str(lead_data.get("telefono", "")).strip() if lead_data.get("telefono") else None,
                "cargo": str(lead_data.get("cargo", "")).strip() if lead_data.get("cargo") else None,
                "sector": str(lead_data.get("sector", "")).strip() if lead_data.get("sector") else None,
                "fuente": str(lead_data.get("fuente", "")).strip() if lead_data.get("fuente") else None,
                "notas": str(lead_data.get("notas", "")).strip() if lead_data.get("notas") else None,
            }
            
            # Handle numeric value
            try:
                valor = lead_data.get("valor_estimado", "0")
                # Remove currency symbols and commas
                if isinstance(valor, str):
                    valor = valor.replace("€", "").replace("$", "").replace(",", ".").strip()
                lead_doc["valor_estimado"] = float(valor) if valor else 0.0
            except:
                lead_doc["valor_estimado"] = 0.0
            
            # Handle etapa
            etapa = str(lead_data.get("etapa", "nuevo")).strip().lower()
            lead_doc["etapa"] = etapa if etapa in LEAD_STAGES else "nuevo"
            
            if action == "update" and existing_id:
                # Update existing lead
                await db.leads.update_one(
                    {"lead_id": existing_id},
                    {"$set": lead_doc}
                )
                updated += 1
            else:
                # Create new lead
                lead_id = f"lead_{uuid.uuid4().hex[:12]}"
                now = datetime.now(timezone.utc)
                lead_doc.update({
                    "lead_id": lead_id,
                    "fecha_creacion": now.isoformat(),
                    "fecha_ultimo_contacto": now.isoformat(),
                    "created_by": current_user.user_id,
                    "servicios": [],
                    "urgencia": "Sin definir",
                })
                await db.leads.insert_one(lead_doc)
                imported += 1
                
        except Exception as e:
            errors.append(f"Fila {idx + 2}: {str(e)}")
    
    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10] if errors else []
    }


@api_router.post("/leads/import")
async def import_leads(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    """Import leads from CSV/Excel (legacy endpoint)"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")
    
    content = await file.read()
    imported = 0
    errors = []
    
    try:
        if file.filename.endswith('.csv'):
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            rows = list(reader)
        elif file.filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Use CSV o Excel.")
        
        for idx, row in enumerate(rows):
            try:
                lead_id = f"lead_{uuid.uuid4().hex[:12]}"
                now = datetime.now(timezone.utc)
                
                lead_doc = {
                    "lead_id": lead_id,
                    "empresa": str(row.get("empresa", "")).strip() or "Sin empresa",
                    "contacto": str(row.get("contacto", "")).strip() or "Sin contacto",
                    "email": str(row.get("email", "")).strip().lower() or f"sin_email_{idx}@temp.com",
                    "telefono": str(row.get("telefono", "")).strip() if row.get("telefono") else None,
                    "cargo": str(row.get("cargo", "")).strip() if row.get("cargo") else None,
                    "sector": str(row.get("sector", "")).strip() if row.get("sector") else None,
                    "valor_estimado": float(row.get("valor_estimado", 0) or 0),
                    "etapa": str(row.get("etapa", "nuevo")).strip().lower() or "nuevo",
                    "notas": str(row.get("notas", "")).strip() if row.get("notas") else None,
                    "fecha_creacion": now.isoformat(),
                    "fecha_ultimo_contacto": now.isoformat(),
                    "created_by": current_user.user_id
                }
                
                if lead_doc["etapa"] not in LEAD_STAGES:
                    lead_doc["etapa"] = "nuevo"
                
                await db.leads.insert_one(lead_doc)
                imported += 1
            except Exception as e:
                errors.append(f"Fila {idx + 2}: {str(e)}")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")
    
    return {
        "message": f"{imported} leads importados",
        "imported": imported,
        "errors": errors[:10] if errors else []
    }

# ============== BULK OPERATIONS ==============

class BulkUpdateRequest(BaseModel):
    lead_ids: List[str]
    etapa: Optional[str] = None
    propietario: Optional[str] = None


@api_router.post("/leads/bulk-update")
async def bulk_update_leads(
    data: BulkUpdateRequest,
    current_user: UserResponse = Depends(get_current_user)
):
    """Bulk update leads stage or owner"""
    update_fields = {}
    
    if data.etapa:
        if data.etapa not in LEAD_STAGES:
            raise HTTPException(status_code=400, detail=f"Etapa inválida")
        update_fields["etapa"] = data.etapa
    
    if data.propietario is not None:
        update_fields["propietario"] = data.propietario if data.propietario else None
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No hay campos para actualizar")
    
    result = await db.leads.update_many(
        {"lead_id": {"$in": data.lead_ids}},
        {"$set": update_fields}
    )
    
    return {"updated": result.modified_count}


class BulkDeleteRequest(BaseModel):
    lead_ids: List[str]


@api_router.post("/leads/bulk-delete")
async def bulk_delete_leads(
    data: BulkDeleteRequest,
    current_user: UserResponse = Depends(get_current_user)
):
    """Bulk delete leads"""
    result = await db.leads.delete_many({"lead_id": {"$in": data.lead_ids}})
    
    # Also delete related activities
    await db.activities.delete_many({"lead_id": {"$in": data.lead_ids}})
    
    return {"deleted": result.deleted_count}


@api_router.delete("/leads/delete-all")
async def delete_all_leads(
    current_user: UserResponse = Depends(get_current_user)
):
    """Delete all leads (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar todos los leads")
    
    result = await db.leads.delete_many({})
    await db.activities.delete_many({})
    
    return {"deleted": result.deleted_count}


# ============== ENRICH API (for Apollo.io) ==============

@api_router.post("/leads/enrich")
async def enrich_lead(data: EnrichRequest):
    """
    API endpoint for external enrichment services like Apollo.io
    Receives email and updates the lead with additional data
    """
    lead = await db.leads.find_one({"email": data.email.lower()}, {"_id": 0})
    
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado con ese email")
    
    update_data = {}
    if data.empresa:
        update_data["empresa"] = data.empresa
    if data.contacto:
        update_data["contacto"] = data.contacto
    if data.telefono:
        update_data["telefono"] = data.telefono
    if data.cargo:
        update_data["cargo"] = data.cargo
    if data.sector:
        update_data["sector"] = data.sector
    
    if update_data:
        await db.leads.update_one(
            {"email": data.email.lower()},
            {"$set": update_data}
        )
    
    return {"message": "Lead enriquecido", "email": data.email, "updated_fields": list(update_data.keys())}

# ============== SECTORS ==============

@api_router.get("/sectors")
async def get_sectors(current_user: UserResponse = Depends(get_current_user)):
    """Get all unique sectors from leads"""
    sectors = await db.leads.distinct("sector")
    return [s for s in sectors if s]

# ============== OPTIONS (Dropdowns) ==============

@api_router.get("/options")
async def get_options(current_user: UserResponse = Depends(get_current_user)):
    """Get all dropdown options"""
    return {
        "sectores": SECTORES,
        "servicios": SERVICIOS,
        "fuentes": FUENTES,
        "urgencias": URGENCIAS,
        "motivos_perdida": MOTIVOS_PERDIDA,
        "tipos_seguimiento": TIPOS_SEGUIMIENTO,
        "etapas": LEAD_STAGES
    }

# ============== REPORTS ==============

@api_router.get("/reports")
async def get_reports(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Get all reports data with optional date filter"""
    
    # Build date filter
    query = {}
    if fecha_inicio and fecha_fin:
        query["fecha_creacion"] = {
            "$gte": fecha_inicio,
            "$lte": fecha_fin + "T23:59:59"
        }
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(10000)
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1}).to_list(100)
    users_map = {u["user_id"]: u["name"] for u in users}
    
    # 1. Pipeline por etapa
    pipeline_por_etapa = {}
    for stage in LEAD_STAGES:
        stage_leads = [l for l in leads if l.get("etapa") == stage]
        pipeline_por_etapa[stage] = {
            "cantidad": len(stage_leads),
            "valor": sum(l.get("valor_estimado", 0) for l in stage_leads)
        }
    
    # 2. Leads por fuente
    leads_por_fuente = {}
    for lead in leads:
        fuente = lead.get("fuente") or "Sin definir"
        leads_por_fuente[fuente] = leads_por_fuente.get(fuente, 0) + 1
    
    # 3. Leads por sector
    leads_por_sector = {}
    for lead in leads:
        sector = lead.get("sector") or "Sin definir"
        leads_por_sector[sector] = leads_por_sector.get(sector, 0) + 1
    
    # 4. Servicios más demandados
    servicios_demandados = {}
    for lead in leads:
        servicios = lead.get("servicios") or []
        for servicio in servicios:
            servicios_demandados[servicio] = servicios_demandados.get(servicio, 0) + 1
    
    # 5. Leads por propietario
    leads_por_propietario = {}
    for lead in leads:
        prop_id = lead.get("propietario")
        prop_name = users_map.get(prop_id, "Sin asignar") if prop_id else "Sin asignar"
        if prop_name not in leads_por_propietario:
            leads_por_propietario[prop_name] = {"cantidad": 0, "valor": 0}
        leads_por_propietario[prop_name]["cantidad"] += 1
        leads_por_propietario[prop_name]["valor"] += lead.get("valor_estimado", 0)
    
    # 6. Motivos de pérdida
    motivos_perdida = {}
    for lead in leads:
        if lead.get("etapa") == "perdido":
            motivo = lead.get("motivo_perdida") or "Sin especificar"
            motivos_perdida[motivo] = motivos_perdida.get(motivo, 0) + 1
    
    return {
        "pipeline_por_etapa": pipeline_por_etapa,
        "leads_por_fuente": leads_por_fuente,
        "leads_por_sector": leads_por_sector,
        "servicios_demandados": servicios_demandados,
        "leads_por_propietario": leads_por_propietario,
        "motivos_perdida": motivos_perdida,
        "total_leads": len(leads)
    }

@api_router.get("/reports/export/{report_type}")
async def export_report(
    report_type: str,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Export report data to CSV"""
    
    # Build date filter
    query = {}
    if fecha_inicio and fecha_fin:
        query["fecha_creacion"] = {
            "$gte": fecha_inicio,
            "$lte": fecha_fin + "T23:59:59"
        }
    
    leads = await db.leads.find(query, {"_id": 0}).to_list(10000)
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1}).to_list(100)
    users_map = {u["user_id"]: u["name"] for u in users}
    
    output = io.StringIO()
    
    if report_type == "pipeline":
        writer = csv.writer(output)
        writer.writerow(["Etapa", "Cantidad", "Valor EUR"])
        for stage in LEAD_STAGES:
            stage_leads = [l for l in leads if l.get("etapa") == stage]
            cantidad = len(stage_leads)
            valor = sum(l.get("valor_estimado", 0) for l in stage_leads)
            writer.writerow([stage.capitalize(), cantidad, valor])
    
    elif report_type == "fuentes":
        writer = csv.writer(output)
        writer.writerow(["Fuente", "Cantidad"])
        fuentes = {}
        for lead in leads:
            fuente = lead.get("fuente") or "Sin definir"
            fuentes[fuente] = fuentes.get(fuente, 0) + 1
        for fuente, cantidad in sorted(fuentes.items(), key=lambda x: x[1], reverse=True):
            writer.writerow([fuente, cantidad])
    
    elif report_type == "sectores":
        writer = csv.writer(output)
        writer.writerow(["Sector", "Cantidad"])
        sectores = {}
        for lead in leads:
            sector = lead.get("sector") or "Sin definir"
            sectores[sector] = sectores.get(sector, 0) + 1
        for sector, cantidad in sorted(sectores.items(), key=lambda x: x[1], reverse=True):
            writer.writerow([sector, cantidad])
    
    elif report_type == "servicios":
        writer = csv.writer(output)
        writer.writerow(["Servicio", "Leads Interesados"])
        servicios = {}
        for lead in leads:
            for servicio in (lead.get("servicios") or []):
                servicios[servicio] = servicios.get(servicio, 0) + 1
        for servicio, cantidad in sorted(servicios.items(), key=lambda x: x[1], reverse=True):
            writer.writerow([servicio, cantidad])
    
    elif report_type == "propietarios":
        writer = csv.writer(output)
        writer.writerow(["Propietario", "Cantidad Leads", "Valor EUR"])
        propietarios = {}
        for lead in leads:
            prop_id = lead.get("propietario")
            prop_name = users_map.get(prop_id, "Sin asignar") if prop_id else "Sin asignar"
            if prop_name not in propietarios:
                propietarios[prop_name] = {"cantidad": 0, "valor": 0}
            propietarios[prop_name]["cantidad"] += 1
            propietarios[prop_name]["valor"] += lead.get("valor_estimado", 0)
        for prop, data in sorted(propietarios.items(), key=lambda x: x[1]["valor"], reverse=True):
            writer.writerow([prop, data["cantidad"], data["valor"]])
    
    elif report_type == "perdidas":
        writer = csv.writer(output)
        writer.writerow(["Motivo", "Cantidad"])
        motivos = {}
        for lead in leads:
            if lead.get("etapa") == "perdido":
                motivo = lead.get("motivo_perdida") or "Sin especificar"
                motivos[motivo] = motivos.get(motivo, 0) + 1
        for motivo, cantidad in sorted(motivos.items(), key=lambda x: x[1], reverse=True):
            writer.writerow([motivo, cantidad])
    
    else:
        raise HTTPException(status_code=400, detail="Tipo de reporte no válido")
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=reporte_{report_type}.csv"}
    )

# ============== USERS (Admin Panel) ==============

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: UserResponse = Depends(get_current_user)):
    """Get all users (for propietario dropdown and admin panel)"""
    users = await db.users.find({}, {"_id": 0}).to_list(100)
    return users

@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: UserResponse = Depends(get_current_user)):
    """Create new user (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden crear usuarios")
    
    # Check if email already exists
    existing = await db.users.find_one({"email": user.email.lower()}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    new_user = {
        "user_id": user_id,
        "email": user.email.lower(),
        "name": user.name,
        "picture": None,
        "role": user.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    return UserResponse(**new_user)

@api_router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(user_id: str, user_update: UserUpdate, current_user: UserResponse = Depends(get_current_user)):
    """Update user (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden editar usuarios")
    
    existing = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    
    if update_data:
        await db.users.update_one({"user_id": user_id}, {"$set": update_data})
    
    updated = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return UserResponse(**updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: UserResponse = Depends(get_current_user)):
    """Delete user (admin only)"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar usuarios")
    
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    
    result = await db.users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    # Also delete user sessions
    await db.user_sessions.delete_many({"user_id": user_id})
    
    return {"message": "Usuario eliminado"}

# ============== OPORTUNIDADES ROUTES ==============

async def get_next_ref_code() -> str:
    """Genera el siguiente ref_code secuencial (01, 02, ... 99, 100, ...)"""
    # Buscar el mayor ref_code existente
    pipeline = [
        {"$match": {"ref_code": {"$exists": True, "$ne": None}}},
        {"$project": {"ref_num": {"$toInt": "$ref_code"}}},
        {"$sort": {"ref_num": -1}},
        {"$limit": 1}
    ]
    result = await db.oportunidades_placsp.aggregate(pipeline).to_list(1)

    if result:
        next_num = result[0]["ref_num"] + 1
    else:
        next_num = 1

    # Formato: "01", "02", ... "99", "100", etc.
    return str(next_num).zfill(2)

@api_router.post("/oportunidades/migrar-ref-codes")
async def migrar_ref_codes(current_user: UserResponse = Depends(get_current_user)):
    """Asigna ref_codes a todas las oportunidades que no tienen uno"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo admin puede ejecutar migraciones")

    # Obtener oportunidades sin ref_code, ordenadas por fecha de detección
    oportunidades = await db.oportunidades_placsp.find(
        {"$or": [{"ref_code": {"$exists": False}}, {"ref_code": None}]},
        {"_id": 1, "oportunidad_id": 1, "fecha_deteccion": 1}
    ).sort("fecha_deteccion", 1).to_list(10000)

    if not oportunidades:
        return {"message": "No hay oportunidades sin ref_code", "migrated": 0}

    # Obtener el último ref_code existente
    next_code = int(await get_next_ref_code())
    migrated = 0

    for op in oportunidades:
        ref_code = str(next_code).zfill(2)
        await db.oportunidades_placsp.update_one(
            {"_id": op["_id"]},
            {"$set": {"ref_code": ref_code}}
        )
        next_code += 1
        migrated += 1

    return {"message": f"Migración completada", "migrated": migrated}

@api_router.get("/oportunidades/tipos-srs")
async def get_tipos_srs():
    """Get available SRS types for filtering - from actual data in DB"""
    # Obtener tipos únicos de la base de datos
    tipos_en_db = await db.oportunidades_placsp.distinct("tipo_srs")
    # Filtrar valores vacíos y ordenar
    tipos_validos = sorted([t for t in tipos_en_db if t and t.strip()])
    return tipos_validos if tipos_validos else TIPOS_SRS

@api_router.get("/oportunidades", response_model=List[OportunidadPLACSP])
async def get_oportunidades(
    tipo_srs: Optional[str] = None,
    score_min: Optional[int] = None,
    convertido_lead: Optional[bool] = None,
    current_user: UserResponse = Depends(get_current_user)
):
    """Get all opportunities with optional filters"""
    query = {}
    if tipo_srs:
        query["tipo_srs"] = tipo_srs
    if score_min is not None:
        query["score"] = {"$gte": score_min}
    if convertido_lead is not None:
        query["convertido_lead"] = convertido_lead
    
    oportunidades = await db.oportunidades_placsp.find(
        query, {"_id": 0}
    ).sort("score", -1).to_list(1000)
    
    # Convert ISO strings back to datetime for response model
    for op in oportunidades:
        for key in ["fecha_adjudicacion", "fecha_fin_contrato", "fecha_deteccion"]:
            if op.get(key) and isinstance(op[key], str):
                try:
                    op[key] = datetime.fromisoformat(op[key].replace("Z", "+00:00"))
                except:
                    pass
    
    return oportunidades

@api_router.post("/oportunidades/spotter")
async def import_oportunidades_spotter(
    data: OportunidadSpotterImport,
    current_user: UserResponse = Depends(get_current_user)
):
    """Import opportunities from SpotterSRS"""
    imported = 0
    duplicates = 0
    
    for oportunidad in data.oportunidades:
        # Check for existing by expediente
        existing = await db.oportunidades_placsp.find_one(
            {"expediente": oportunidad.expediente}
        )
        
        if existing:
            duplicates += 1
            continue
        
        oportunidad_id = f"op_{uuid.uuid4().hex[:12]}"
        oportunidad_dict = oportunidad.model_dump()
        
        # Convert datetime to ISO strings for MongoDB
        for key in ["fecha_adjudicacion", "fecha_fin_contrato", "fecha_deteccion"]:
            if oportunidad_dict.get(key):
                if isinstance(oportunidad_dict[key], datetime):
                    oportunidad_dict[key] = oportunidad_dict[key].isoformat()
        
        # Generar ref_code único
        ref_code = await get_next_ref_code()

        oportunidad_doc = {
            "oportunidad_id": oportunidad_id,
            "ref_code": ref_code,
            **oportunidad_dict
        }

        await db.oportunidades_placsp.insert_one(oportunidad_doc)
        imported += 1

    return {
        "message": "Importación completada",
        "imported": imported,
        "duplicates": duplicates,
        "total": len(data.oportunidades)
    }


@api_router.post("/oportunidades/spotter-internal")
async def import_oportunidades_spotter_internal(data: OportunidadSpotterImport):
    """Internal endpoint for SpotterSRS cron - no auth required, localhost only"""
    imported = 0
    duplicates = 0
    for oportunidad in data.oportunidades:
        existing = await db.oportunidades_placsp.find_one({"expediente": oportunidad.expediente})
        if existing:
            duplicates += 1
            continue
        oportunidad_id = f"op_{uuid.uuid4().hex[:12]}"
        oportunidad_dict = oportunidad.model_dump()
        for key in ["fecha_adjudicacion", "fecha_fin_contrato", "fecha_deteccion"]:
            if oportunidad_dict.get(key):
                if isinstance(oportunidad_dict[key], datetime):
                    oportunidad_dict[key] = oportunidad_dict[key].isoformat()
        # Generar ref_code único
        ref_code = await get_next_ref_code()
        oportunidad_doc = {"oportunidad_id": oportunidad_id, "ref_code": ref_code, **oportunidad_dict}
        await db.oportunidades_placsp.insert_one(oportunidad_doc)
        imported += 1
    return {"message": "Importacion completada", "imported": imported, "duplicates": duplicates, "total": len(data.oportunidades)}


@api_router.post("/oportunidades/ejecutar-spotter")
async def ejecutar_spotter_manual(current_user: UserResponse = Depends(get_current_user)):
    """
    Ejecuta SpotterSRS manualmente para buscar nuevas oportunidades.
    Solo usuarios autenticados pueden ejecutar esta acción.
    """
    import subprocess
    import os as os_module

    # Ruta al script de SpotterSRS
    spotter_dir = os_module.path.join(os_module.path.dirname(__file__), "app", "spotter")
    spotter_script = os_module.path.join(spotter_dir, "run_spotter_cron.py")

    if not os_module.path.exists(spotter_script):
        raise HTTPException(
            status_code=500,
            detail=f"Script SpotterSRS no encontrado en {spotter_script}"
        )

    try:
        # Ejecutar el script de SpotterSRS
        result = subprocess.run(
            ["python3", spotter_script],
            cwd=spotter_dir,
            capture_output=True,
            text=True,
            timeout=300  # 5 minutos máximo
        )

        if result.returncode == 0:
            # Contar oportunidades nuevas (parsear output si es posible)
            output = result.stdout

            return {
                "success": True,
                "message": "SpotterSRS ejecutado correctamente",
                "output": output[-2000:] if len(output) > 2000 else output,  # Últimos 2000 chars
                "executed_by": current_user.email,
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
        else:
            return {
                "success": False,
                "message": "SpotterSRS terminó con errores",
                "error": result.stderr[-1000:] if result.stderr else "Sin detalles",
                "executed_by": current_user.email,
                "timestamp": datetime.now(timezone.utc).isoformat()
            }

    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=504,
            detail="SpotterSRS excedió el tiempo límite de 5 minutos"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error ejecutando SpotterSRS: {str(e)}"
        )


@api_router.post("/oportunidades/reclasificar")
async def reclasificar_oportunidades_endpoint(current_user: UserResponse = Depends(get_current_user)):
    """
    Reclasifica todas las oportunidades existentes usando el algoritmo actualizado.
    Solo usuarios autenticados pueden ejecutar esta acción.
    """
    import sys
    import os as os_module

    # Añadir el directorio de spotter al path
    spotter_dir = os_module.path.join(os_module.path.dirname(__file__), "app", "spotter")
    if spotter_dir not in sys.path:
        sys.path.insert(0, spotter_dir)

    try:
        from app.spotter.spotter_srs import calcular_dolor, extraer_keywords

        # Obtener todas las oportunidades
        oportunidades = await db.oportunidades_placsp.find(
            {},
            {"_id": 0}
        ).to_list(10000)

        if not oportunidades:
            return {
                "success": True,
                "message": "No hay oportunidades para reclasificar",
                "total": 0,
                "cambios": 0
            }

        cambios = 0
        errores = 0

        for opp in oportunidades:
            try:
                oportunidad_id = opp.get('oportunidad_id')
                objeto = opp.get('objeto', '')
                cpv = opp.get('cpv', '')
                tipo_actual = opp.get('tipo_srs', '')
                fecha_adj = opp.get('fecha_adjudicacion', '')
                dias_restantes = opp.get('dias_restantes')

                # Extraer keywords
                keywords = extraer_keywords(objeto)

                # Convertir fecha
                if isinstance(fecha_adj, datetime):
                    fecha_adj_str = fecha_adj.strftime('%Y-%m-%d')
                else:
                    fecha_adj_str = str(fecha_adj)[:10] if fecha_adj else ''

                # Calcular nueva clasificación
                dolor = calcular_dolor(
                    objeto=objeto,
                    fecha_adjudicacion=fecha_adj_str,
                    duracion_dias=dias_restantes,
                    cpv=cpv,
                    keywords=keywords
                )

                nuevo_tipo = dolor.tipo_oportunidad.value

                # Actualizar si cambió
                if tipo_actual != nuevo_tipo:
                    await db.oportunidades_placsp.update_one(
                        {"oportunidad_id": oportunidad_id},
                        {"$set": {
                            "tipo_srs": nuevo_tipo,
                            "keywords": list(keywords.keys()),
                            "indicadores_dolor": dolor.indicadores_urgencia,
                            "score": dolor.score_dolor,
                        }}
                    )
                    cambios += 1

            except Exception as e:
                errores += 1
                print(f"Error reclasificando {opp.get('oportunidad_id')}: {e}")

        return {
            "success": True,
            "message": "Reclasificación completada",
            "total": len(oportunidades),
            "cambios": cambios,
            "sin_cambios": len(oportunidades) - cambios - errores,
            "errores": errores,
            "executed_by": current_user.email,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando módulo SpotterSRS: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en reclasificación: {str(e)}"
        )


@api_router.post("/oportunidades/{oportunidad_id}/analizar-pain")
async def analizar_pain_oportunidad_endpoint(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Analiza el pain/urgencia de una oportunidad usando IA multi-proveedor.
    Orden: OpenAI -> Gemini -> Claude -> Básico
    Respuesta esperada: <5 segundos
    """
    try:
        from app.spotter.pain_analyzer import analizar_pain_oportunidad

        # Obtener la oportunidad con importe
        oportunidad = await db.oportunidades_placsp.find_one(
            {"oportunidad_id": oportunidad_id},
            {"_id": 0}
        )

        if not oportunidad:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

        # Analizar pain (sin PDFs para velocidad)
        resultado = await analizar_pain_oportunidad(
            oportunidad_id=oportunidad_id,
            objeto=oportunidad.get("objeto", ""),
            cpv=oportunidad.get("cpv", ""),
            importe=float(oportunidad.get("importe", 0) or 0)
        )

        # Guardar análisis en la oportunidad
        await db.oportunidades_placsp.update_one(
            {"oportunidad_id": oportunidad_id},
            {"$set": {
                "pain_analysis": resultado,
                "pain_score": resultado.get("pain_score", 0),
                "nivel_urgencia": resultado.get("nivel_urgencia", "medio"),
                "analisis_fecha": datetime.now(timezone.utc).isoformat()
            }}
        )

        return {
            "success": True,
            "oportunidad_id": oportunidad_id,
            "pain_analysis": resultado
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando módulo pain_analyzer: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error analizando pain: {str(e)}"
        )


@api_router.post("/oportunidades/analizar-pain-batch")
async def analizar_pain_batch(
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Analiza el pain de oportunidades sin análisis.
    Procesa 10 oportunidades por batch (rápido sin PDFs).
    """
    try:
        from app.spotter.pain_analyzer import analizar_pain_oportunidad

        # Obtener oportunidades sin análisis de pain
        oportunidades = await db.oportunidades_placsp.find(
            {"pain_analysis": {"$exists": False}},
            {"_id": 0, "oportunidad_id": 1, "objeto": 1, "cpv": 1, "importe": 1}
        ).limit(10).to_list(10)

        if not oportunidades:
            return {
                "success": True,
                "message": "No hay oportunidades pendientes de análisis",
                "analizadas": 0
            }

        analizadas = 0
        errores = 0
        detalles = []

        for opp in oportunidades:
            opp_id = opp["oportunidad_id"]
            try:
                # Análisis rápido (sin PDFs, <5s por oportunidad)
                resultado = await analizar_pain_oportunidad(
                    oportunidad_id=opp_id,
                    objeto=opp.get("objeto", ""),
                    cpv=opp.get("cpv", ""),
                    importe=float(opp.get("importe", 0) or 0)
                )

                await db.oportunidades_placsp.update_one(
                    {"oportunidad_id": opp_id},
                    {"$set": {
                        "pain_analysis": resultado,
                        "pain_score": resultado.get("pain_score", 0),
                        "nivel_urgencia": resultado.get("nivel_urgencia", "medio"),
                        "analisis_fecha": datetime.now(timezone.utc).isoformat()
                    }}
                )
                analizadas += 1
                detalles.append({
                    "id": opp_id,
                    "status": "ok",
                    "score": resultado.get("pain_score", 0),
                    "proveedor": resultado.get("proveedor_ia", "unknown")
                })

            except Exception as e:
                errores += 1
                detalles.append({"id": opp_id, "status": "error", "error": str(e)[:100]})
                print(f"Error analizando {opp_id}: {e}")

        pendientes = await db.oportunidades_placsp.count_documents(
            {"pain_analysis": {"$exists": False}}
        )

        return {
            "success": True,
            "message": f"Análisis completado",
            "analizadas": analizadas,
            "errores": errores,
            "pendientes": pendientes,
            "detalles": detalles
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando módulo pain_analyzer: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en análisis batch: {str(e)}"
        )


@api_router.post("/oportunidades/{oportunidad_id}/analizar-pliego")
async def analizar_pliego_exhaustivo(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Análisis EXHAUSTIVO del pliego de una oportunidad.
    Descarga el PDF/HTML, extrae texto, analiza con IA buscando IT oculto.
    Genera resumen orientado al operador comercial.

    Timeout máximo: 150 segundos.
    """
    import asyncio
    import logging
    logger = logging.getLogger(__name__)

    try:
        from app.spotter.pliego_analyzer import analizar_pliego_completo

        # Buscar la oportunidad
        oportunidad = await db.oportunidades_placsp.find_one(
            {"oportunidad_id": oportunidad_id}
        )

        if not oportunidad:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

        # Buscar URL del pliego
        url_pliego = None
        pliegos = oportunidad.get("pliegos", {})
        datos_adj = oportunidad.get("datos_adjudicatario", {})

        # Priorizar pliego técnico, luego administrativo
        # Buscar en pliegos y también en datos_adjudicatario (donde lo guarda el enricher)
        if pliegos.get("url_pliego_tecnico"):
            url_pliego = pliegos["url_pliego_tecnico"]
        elif datos_adj.get("url_pliego_tecnico"):
            url_pliego = datos_adj["url_pliego_tecnico"]
        elif pliegos.get("url_pliego_administrativo"):
            url_pliego = pliegos["url_pliego_administrativo"]
        elif datos_adj.get("url_pliego_administrativo"):
            url_pliego = datos_adj["url_pliego_administrativo"]

        # Si no hay URL de pliego, intentar extraerla de la página de PLACSP
        if not url_pliego and oportunidad.get("url_licitacion"):
            logger.info(f"No hay URL de pliego guardada, extrayendo de PLACSP...")
            try:
                from app.spotter.adjudicatario_enricher import AdjudicatarioEnricher
                enricher = AdjudicatarioEnricher()

                # Extraer datos de la página de licitación (incluye URLs de pliegos)
                datos_extraidos = await enricher.extraer_datos_placsp(
                    url_licitacion=oportunidad["url_licitacion"],
                    nombre_adjudicatario=oportunidad.get("adjudicatario", ""),
                    nif_adjudicatario=oportunidad.get("nif", "")
                )

                # Buscar URL de pliego en los datos extraídos
                if datos_extraidos.get("url_pliego_tecnico"):
                    url_pliego = datos_extraidos["url_pliego_tecnico"]
                    logger.info(f"✓ PPT extraído de PLACSP: {url_pliego[:80]}")
                elif datos_extraidos.get("url_pliego_administrativo"):
                    url_pliego = datos_extraidos["url_pliego_administrativo"]
                    logger.info(f"✓ PCAP extraído de PLACSP: {url_pliego[:80]}")

                # Guardar datos extraídos para futuras consultas
                if datos_extraidos:
                    await db.oportunidades_placsp.update_one(
                        {"oportunidad_id": oportunidad_id},
                        {"$set": {"datos_adjudicatario": datos_extraidos}}
                    )
                    logger.info(f"✓ Datos adjudicatario guardados con URLs de pliegos")
            except Exception as e:
                logger.warning(f"Error extrayendo datos de PLACSP: {e}")

        # Último fallback: usar URL de licitación (analizará la página HTML)
        if not url_pliego and oportunidad.get("url_licitacion"):
            url_pliego = oportunidad["url_licitacion"]
            logger.warning(f"Usando URL de licitación como fallback (no es el PDF del pliego)")

        if not url_pliego:
            raise HTTPException(
                status_code=400,
                detail="No se encontró URL de pliego para esta oportunidad"
            )

        logger.info(f"Iniciando análisis de pliego para {oportunidad_id}, URL: {url_pliego}")

        # Ejecutar análisis con timeout global de 150 segundos
        try:
            resultado = await asyncio.wait_for(
                analizar_pliego_completo(
                    oportunidad_id=oportunidad_id,
                    url_pliego=url_pliego,
                    objeto=oportunidad.get("objeto", ""),
                    importe=oportunidad.get("importe", 0)
                ),
                timeout=150.0  # 2.5 minutos máximo
            )
        except asyncio.TimeoutError:
            logger.error(f"Timeout de 150s en análisis de pliego para {oportunidad_id}")
            raise HTTPException(
                status_code=504,
                detail="El análisis tardó demasiado (>150s). Intente de nuevo más tarde."
            )

        # Guardar resultado en BD
        await db.oportunidades_placsp.update_one(
            {"oportunidad_id": oportunidad_id},
            {
                "$set": {
                    "analisis_pliego": resultado,
                    "analisis_pliego_fecha": datetime.now(timezone.utc).isoformat(),
                    "tiene_it_confirmado": resultado.get("tiene_it", False),
                    "pain_score": resultado.get("pain_score", 0),
                    "nivel_urgencia": resultado.get("nivel_urgencia", "bajo")
                }
            }
        )

        return {
            "success": True,
            "oportunidad_id": oportunidad_id,
            "url_pliego_analizado": url_pliego,
            "tiempo_analisis": resultado.get("tiempo_analisis_segundos", 0),
            "analisis": resultado
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando módulo pliego_analyzer: {str(e)}"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error en análisis de pliego: {str(e)}"
        )


@api_router.get("/oportunidades/{oportunidad_id}/resumen-operador")
async def obtener_resumen_operador(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Obtiene el resumen orientado al OPERADOR de una oportunidad.
    Ideal para preparar correos/llamadas con el dolor específico del cliente.

    Requiere que se haya ejecutado primero /analizar-pliego
    """
    oportunidad = await db.oportunidades_placsp.find_one(
        {"oportunidad_id": oportunidad_id},
        {"analisis_pliego": 1, "objeto": 1, "adjudicatario": 1, "importe": 1, "nif": 1, "organo_contratacion": 1, "datos_adjudicatario": 1}
    )

    if not oportunidad:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

    analisis = oportunidad.get("analisis_pliego")

    if not analisis:
        raise HTTPException(
            status_code=400,
            detail="Esta oportunidad no tiene análisis de pliego. Ejecute primero POST /analizar-pliego"
        )

    resumen = analisis.get("resumen_operador", {})

    return {
        "oportunidad_id": oportunidad_id,
        "empresa": oportunidad.get("adjudicatario", ""),
        "objeto": oportunidad.get("objeto", ""),
        "importe": oportunidad.get("importe", 0),
        "nif": oportunidad.get("nif", ""),
        "organo_contratacion": oportunidad.get("organo_contratacion", ""),
        "adjudicatario": oportunidad.get("adjudicatario", ""),
        "datos_adjudicatario": oportunidad.get("datos_adjudicatario"),
        "tiene_it": resumen.get("tiene_it", False),
        "nivel_oportunidad": resumen.get("nivel_oportunidad", ""),
        "dolor_principal": resumen.get("dolor_principal", ""),
        "gancho_email": resumen.get("gancho_inicial", ""),
        "puntos_dolor": resumen.get("puntos_dolor_email", []),
        "preguntas_cualificacion": resumen.get("preguntas_cualificacion", []),
        "tecnologias": resumen.get("tecnologias_mencionadas", []),
        "certificaciones": resumen.get("certificaciones_requeridas", []),
        "alertas": resumen.get("alertas", []),
        "confianza": resumen.get("confianza_analisis", "")
    }


@api_router.post("/oportunidades/{oportunidad_id}/analisis-comercial")
async def generar_analisis_comercial_endpoint(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Genera un ANÁLISIS COMERCIAL COMPLETO orientado a la acción.

    Incluye:
    - Clasificación y scoring detallado
    - Resumen ejecutivo (30 segundos)
    - Información del adjudicatario
    - Dolores principales y secundarios
    - Servicios SRS aplicables con valor estimado
    - Competencia y ventajas
    - Comunicación lista para usar (WhatsApp, email, llamada)
    - Objeciones y respuestas preparadas
    - Siguientes pasos

    Timeout máximo: 180 segundos.
    """
    import asyncio
    import logging
    logger = logging.getLogger(__name__)

    try:
        from app.spotter.pliego_analyzer import generar_analisis_comercial

        # Buscar la oportunidad con todos sus datos
        oportunidad = await db.oportunidades_placsp.find_one(
            {"oportunidad_id": oportunidad_id}
        )

        if not oportunidad:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

        # Buscar URL del pliego
        url_pliego = None
        pliegos = oportunidad.get("pliegos", {})
        datos_adj = oportunidad.get("datos_adjudicatario", {})

        # Priorizar pliego técnico, buscar en pliegos y datos_adjudicatario
        if pliegos.get("url_pliego_tecnico"):
            url_pliego = pliegos["url_pliego_tecnico"]
        elif datos_adj.get("url_pliego_tecnico"):
            url_pliego = datos_adj["url_pliego_tecnico"]
        elif pliegos.get("url_pliego_admin"):
            url_pliego = pliegos["url_pliego_admin"]
        elif datos_adj.get("url_pliego_administrativo"):
            url_pliego = datos_adj["url_pliego_administrativo"]
        elif oportunidad.get("url_licitacion"):
            url_pliego = oportunidad["url_licitacion"]

        if not url_pliego:
            raise HTTPException(
                status_code=400,
                detail="Esta oportunidad no tiene URL de pliego disponible"
            )

        logger.info(f"Generando análisis comercial para {oportunidad_id}")

        # Datos del adjudicatario
        datos_adj = oportunidad.get("datos_adjudicatario", {})

        # Ejecutar análisis con timeout
        try:
            resultado = await asyncio.wait_for(
                generar_analisis_comercial(
                    oportunidad_id=oportunidad_id,
                    url_pliego=url_pliego,
                    objeto=oportunidad.get("objeto", ""),
                    importe=oportunidad.get("importe", 0),
                    adjudicatario_nombre=oportunidad.get("adjudicatario", ""),
                    adjudicatario_cif=oportunidad.get("nif", ""),
                    organo_contratante=oportunidad.get("organo_contratacion", ""),
                    fecha_adjudicacion=oportunidad.get("fecha_adjudicacion", ""),
                ),
                timeout=180.0  # 3 minutos máximo
            )
        except asyncio.TimeoutError:
            logger.error(f"Timeout de 180s en análisis comercial para {oportunidad_id}")
            raise HTTPException(
                status_code=504,
                detail="El análisis tardó demasiado (>180s). Intente de nuevo más tarde."
            )

        if resultado.get("error"):
            raise HTTPException(
                status_code=500,
                detail=resultado["error"]
            )

        # Guardar resultado en BD
        await db.oportunidades_placsp.update_one(
            {"oportunidad_id": oportunidad_id},
            {
                "$set": {
                    "analisis_comercial": resultado,
                    "analisis_comercial_fecha": datetime.now()
                }
            }
        )

        return {
            "success": True,
            "oportunidad_id": oportunidad_id,
            "analisis_comercial": resultado
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en análisis comercial: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error en análisis comercial: {str(e)}"
        )


@api_router.get("/oportunidades/{oportunidad_id}/analisis-comercial")
async def obtener_analisis_comercial(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Obtiene el análisis comercial completo de una oportunidad.
    Requiere que se haya generado previamente con POST /analisis-comercial
    """
    oportunidad = await db.oportunidades_placsp.find_one(
        {"oportunidad_id": oportunidad_id},
        {"analisis_comercial": 1, "analisis_comercial_fecha": 1}
    )

    if not oportunidad:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

    analisis = oportunidad.get("analisis_comercial")

    if not analisis:
        raise HTTPException(
            status_code=400,
            detail="Esta oportunidad no tiene análisis comercial. Ejecute primero POST /analisis-comercial"
        )

    return {
        "oportunidad_id": oportunidad_id,
        "fecha_analisis": oportunidad.get("analisis_comercial_fecha"),
        "analisis_comercial": analisis
    }


@api_router.post("/oportunidades/{oportunidad_id}/enriquecer-adjudicatario")
async def enriquecer_adjudicatario_endpoint(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Enriquece los datos del adjudicatario desde fuentes externas.

    Busca información adicional como:
    - Dirección completa
    - Teléfono de contacto
    - Email de contacto
    - Sitio web
    - Actividad/CNAE
    - Número de empleados

    Fuentes: Infocif.es, Einforma.com, PLACSP
    """
    try:
        from app.spotter.adjudicatario_enricher import enriquecer_adjudicatario

        # Obtener la oportunidad
        oportunidad = await db.oportunidades_placsp.find_one(
            {"oportunidad_id": oportunidad_id},
            {"_id": 0, "adjudicatario": 1, "nif": 1, "url_licitacion": 1, "datos_adjudicatario": 1}
        )

        if not oportunidad:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

        nombre = oportunidad.get("adjudicatario", "")
        nif = oportunidad.get("nif", "")
        url_licitacion = oportunidad.get("url_licitacion")

        if not nif:
            raise HTTPException(status_code=400, detail="La oportunidad no tiene NIF del adjudicatario")

        # Enriquecer datos
        datos = await enriquecer_adjudicatario(
            nombre=nombre,
            nif=nif,
            url_licitacion=url_licitacion
        )

        # Guardar en la base de datos
        await db.oportunidades_placsp.update_one(
            {"oportunidad_id": oportunidad_id},
            {"$set": {"datos_adjudicatario": datos}}
        )

        return {
            "success": True,
            "oportunidad_id": oportunidad_id,
            "datos_adjudicatario": datos
        }

    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error importando módulo adjudicatario_enricher: {str(e)}"
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error enriqueciendo adjudicatario: {str(e)}"
        )


class EstadoRevisionUpdate(BaseModel):
    estado: str  # nueva, revisada, descartada


@api_router.patch("/oportunidades/{oportunidad_id}/estado-revision")
async def update_estado_revision(
    oportunidad_id: str,
    data: EstadoRevisionUpdate,
    current_user: UserResponse = Depends(get_current_user)
):
    """Actualizar estado de revisión de una oportunidad"""
    estados_validos = ["nueva", "revisada", "descartada"]
    if data.estado not in estados_validos:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Use: {estados_validos}")

    result = await db.oportunidades_placsp.update_one(
        {"oportunidad_id": oportunidad_id},
        {"$set": {
            "estado_revision": data.estado,
            "fecha_revision": datetime.now(timezone.utc).isoformat(),
            "revisado_por": current_user.user_id
        }}
    )

    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

    return {
        "success": True,
        "oportunidad_id": oportunidad_id,
        "estado_revision": data.estado
    }


@api_router.post("/oportunidades/{oportunidad_id}/convertir-lead")
async def convert_oportunidad_to_lead(
    oportunidad_id: str,
    current_user: UserResponse = Depends(get_current_user)
):
    """Convert opportunity to lead, incluyendo datos del análisis de pliego"""
    oportunidad = await db.oportunidades_placsp.find_one(
        {"oportunidad_id": oportunidad_id}, {"_id": 0}
    )

    if not oportunidad:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

    if oportunidad.get("convertido_lead"):
        raise HTTPException(status_code=400, detail="Ya convertido a lead")

    # Create lead from oportunidad
    lead_id = f"lead_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)

    # Extraer datos del análisis de pliego si existe
    analisis_pliego = oportunidad.get("analisis_pliego", {})
    resumen_operador = analisis_pliego.get("resumen_operador", {})

    # Construir notas enriquecidas
    notas_base = f"Origen: PLACSP - {oportunidad['expediente']}\n"
    notas_base += f"Objeto: {oportunidad['objeto']}\n"
    notas_base += f"Score: {oportunidad['score']}\n"
    notas_base += f"CPV: {oportunidad['cpv']}\n"
    notas_base += f"Órgano: {oportunidad['organo_contratacion']}\n"
    notas_base += f"URL: {oportunidad['url_licitacion']}\n"

    # Añadir info del análisis de pliego
    if resumen_operador:
        notas_base += f"\n--- ANÁLISIS IA ---\n"
        if resumen_operador.get("dolor_principal"):
            notas_base += f"Dolor principal: {resumen_operador['dolor_principal']}\n"
        if resumen_operador.get("gancho_inicial"):
            notas_base += f"Gancho: {resumen_operador['gancho_inicial']}\n"
        if resumen_operador.get("nivel_oportunidad"):
            notas_base += f"Nivel: {resumen_operador['nivel_oportunidad'].upper()}\n"
        if resumen_operador.get("puntos_dolor_email"):
            notas_base += f"Puntos dolor: {', '.join(resumen_operador['puntos_dolor_email'][:3])}\n"
        if resumen_operador.get("tecnologias_mencionadas"):
            notas_base += f"Tecnologías: {', '.join(resumen_operador['tecnologias_mencionadas'][:5])}\n"
        if resumen_operador.get("certificaciones_requeridas"):
            notas_base += f"Certificaciones: {', '.join(resumen_operador['certificaciones_requeridas'])}\n"

    # Determinar urgencia basada en días restantes
    dias_restantes = oportunidad.get("dias_restantes")
    urgencia = "Sin definir"
    if dias_restantes is not None:
        if dias_restantes < 30:
            urgencia = "Inmediata (< 1 mes)"
        elif dias_restantes < 90:
            urgencia = "Corto plazo (1-3 meses)"
        elif dias_restantes < 180:
            urgencia = "Medio plazo (3-6 meses)"
        else:
            urgencia = "Largo plazo (6+ meses)"

    lead_doc = {
        "lead_id": lead_id,
        "empresa": oportunidad["adjudicatario"],
        "contacto": "",  # To be filled manually
        "email": "",     # To be filled manually
        "telefono": None,
        "cargo": None,
        "sector": oportunidad.get("tipo_srs", "Otro"),
        "valor_estimado": oportunidad.get("importe", 0),
        "etapa": "nuevo",
        "notas": notas_base,
        "propietario": current_user.user_id,
        "servicios": [],
        "fuente": "Licitación",
        "urgencia": urgencia,
        "fecha_creacion": now.isoformat(),
        "fecha_ultimo_contacto": now.isoformat(),
        "created_by": current_user.user_id,
        # Datos extra del pliego para referencia
        "oportunidad_origen": oportunidad_id,
        "nivel_oportunidad": resumen_operador.get("nivel_oportunidad"),
        "dolor_principal": resumen_operador.get("dolor_principal"),
        "gancho_inicial": resumen_operador.get("gancho_inicial")
    }

    await db.leads.insert_one(lead_doc)

    # Mark oportunidad as converted
    await db.oportunidades_placsp.update_one(
        {"oportunidad_id": oportunidad_id},
        {"$set": {"convertido_lead": True}}
    )

    return {
        "message": "Oportunidad convertida a lead",
        "lead_id": lead_id,
        "oportunidad_id": oportunidad_id,
        "datos_pliego_incluidos": bool(resumen_operador)
    }

# ============== ROOT ==============

# ============== APOLLO.IO INTEGRATION ==============

APOLLO_API_KEY = os.environ.get("APOLLO_API_KEY", "")
APOLLO_BASE_URL = "https://api.apollo.io/v1"


class ApolloEnrichPersonRequest(BaseModel):
    email: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    domain: Optional[str] = None


class ApolloEnrichCompanyRequest(BaseModel):
    domain: str


class ApolloSearchRequest(BaseModel):
    job_titles: Optional[List[str]] = None
    locations: Optional[List[str]] = None
    domains: Optional[List[str]] = None
    industries: Optional[List[str]] = None
    company_sizes: Optional[List[str]] = None
    page: int = 1
    per_page: int = 25


@api_router.post("/apollo/enrich/person")
async def apollo_enrich_person(request: ApolloEnrichPersonRequest, current_user: UserResponse = Depends(get_current_user)):
    if not APOLLO_API_KEY:
        raise HTTPException(status_code=500, detail="Apollo API key not configured")
    if not request.email and not (request.first_name and request.last_name and request.domain):
        raise HTTPException(status_code=400, detail="Provide email OR (first_name + last_name + domain)")
    headers = {"Content-Type": "application/json", "Cache-Control": "no-cache", "X-Api-Key": APOLLO_API_KEY}
    payload = {}
    if request.email:
        payload["email"] = request.email
    if request.first_name:
        payload["first_name"] = request.first_name
    if request.last_name:
        payload["last_name"] = request.last_name
    if request.domain:
        payload["organization_domain"] = request.domain
    async with httpx.AsyncClient() as client:
        response = await client.post(f"{APOLLO_BASE_URL}/people/match", headers=headers, json=payload, timeout=30.0)
        if response.status_code == 200:
            data = response.json()
            person = data.get("person", {})
            return {"success": True, "data": {"first_name": person.get("first_name"), "last_name": person.get("last_name"), "email": person.get("email"), "title": person.get("title"), "linkedin_url": person.get("linkedin_url"), "phone": person.get("phone_numbers", [{}])[0].get("raw_number") if person.get("phone_numbers") else None, "company": person.get("organization", {}).get("name"), "company_domain": person.get("organization", {}).get("primary_domain"), "company_industry": person.get("organization", {}).get("industry"), "company_size": person.get("organization", {}).get("estimated_num_employees"), "city": person.get("city"), "country": person.get("country")}}
        else:
            raise HTTPException(status_code=502, detail=f"Apollo API error: {response.status_code}")


@api_router.post("/apollo/enrich/company")
async def apollo_enrich_company(request: ApolloEnrichCompanyRequest, current_user: UserResponse = Depends(get_current_user)):
    if not APOLLO_API_KEY:
        raise HTTPException(status_code=500, detail="Apollo API key not configured")
    headers = {"Content-Type": "application/json", "Cache-Control": "no-cache", "X-Api-Key": APOLLO_API_KEY}
    async with httpx.AsyncClient() as client:
        response = await client.get(f"{APOLLO_BASE_URL}/organizations/enrich", headers=headers, params={"domain": request.domain}, timeout=30.0)
        if response.status_code == 200:
            data = response.json()
            org = data.get("organization", {})
            return {"success": True, "data": {"name": org.get("name"), "domain": org.get("primary_domain"), "industry": org.get("industry"), "estimated_employees": org.get("estimated_num_employees"), "linkedin_url": org.get("linkedin_url"), "website_url": org.get("website_url"), "phone": org.get("phone"), "city": org.get("city"), "country": org.get("country"), "description": org.get("short_description"), "founded_year": org.get("founded_year"), "annual_revenue": org.get("annual_revenue_printed")}}
        else:
            raise HTTPException(status_code=502, detail=f"Apollo API error: {response.status_code}")


@api_router.post("/apollo/search")
async def apollo_search_prospects(request: ApolloSearchRequest, current_user: UserResponse = Depends(get_current_user)):
    if not APOLLO_API_KEY:
        raise HTTPException(status_code=500, detail="Apollo API key not configured")
    headers = {"Content-Type": "application/json", "Cache-Control": "no-cache", "X-Api-Key": APOLLO_API_KEY}
    payload = {"page": request.page, "per_page": request.per_page}
    if request.job_titles:
        payload["person_titles"] = request.job_titles
    if request.locations:
        payload["person_locations"] = request.locations
    if request.domains:
        payload["organization_domains"] = request.domains
    if request.industries:
        payload["organization_industry_tag_ids"] = request.industries
    if request.company_sizes:
        payload["organization_num_employees_ranges"] = request.company_sizes
    async with httpx.AsyncClient() as client:
        response = await client.post(f"{APOLLO_BASE_URL}/mixed_people/search", headers=headers, json=payload, timeout=30.0)
        if response.status_code == 200:
            data = response.json()
            people = data.get("people", [])
            return {"success": True, "total": data.get("pagination", {}).get("total_entries", 0), "page": request.page, "per_page": request.per_page, "data": [{"id": p.get("id"), "first_name": p.get("first_name"), "last_name": p.get("last_name"), "email": p.get("email"), "title": p.get("title"), "linkedin_url": p.get("linkedin_url"), "company": p.get("organization", {}).get("name") if p.get("organization") else None, "company_domain": p.get("organization", {}).get("primary_domain") if p.get("organization") else None, "city": p.get("city"), "country": p.get("country")} for p in people]}
        else:
            raise HTTPException(status_code=502, detail=f"Apollo API error: {response.status_code}")


@api_router.post("/leads/{lead_id}/enrich-apollo")
async def enrich_lead_with_apollo(lead_id: str, current_user: UserResponse = Depends(get_current_user)):
    if not APOLLO_API_KEY:
        raise HTTPException(status_code=500, detail="Apollo API key not configured")
    lead = await db.leads.find_one({"lead_id": lead_id}, {"_id": 0})
    if not lead:
        raise HTTPException(status_code=404, detail="Lead no encontrado")
    
    # Extraer dominio del email o generar desde empresa
    domain = None
    if lead.get("email") and "@" in lead.get("email", ""):
        domain = lead["email"].split("@")[1]
    elif lead.get("empresa"):
        empresa = lead["empresa"].lower()
        # Limpiar sufijos legales
        for suffix in ["sociedad limitada", "sociedad anonima", "s.l.", "s.a.", "sl", "sa", "s.l.u.", "slu", "s.a.u.", "sau", "limited", "ltd", "inc", "corp"]:
            empresa = empresa.replace(suffix, "")
        empresa = empresa.replace(" ", "").replace(",", "").replace(".", "").strip()
        domain = f"{empresa}.com"
    
    if not domain:
        raise HTTPException(status_code=400, detail="No hay suficiente info para enriquecer (necesita email o empresa)")
    
    headers = {"Content-Type": "application/json", "X-Api-Key": APOLLO_API_KEY}
    
    async with httpx.AsyncClient() as client:
        # Enriquecer organización
        response = await client.post(
            f"{APOLLO_BASE_URL}/organizations/enrich",
            headers=headers,
            json={"domain": domain},
            timeout=30.0
        )
        
        if response.status_code == 200:
            data = response.json()
            org = data.get("organization", {})
            
            if not org:
                return {"success": False, "message": "No se encontró info en Apollo"}
            
            update_data = {}
            
            # Actualizar teléfono si no existe
            if org.get("phone") and not lead.get("telefono"):
                update_data["telefono"] = org["phone"]
            
            # Actualizar sector si no existe
            if org.get("industry") and not lead.get("sector"):
                update_data["sector"] = org["industry"].title()
            
            # Añadir info a notas
            notas_extra = []
            if org.get("linkedin_url"):
                notas_extra.append(f"LinkedIn: {org['linkedin_url']}")
            if org.get("estimated_num_employees"):
                notas_extra.append(f"Empleados: {org['estimated_num_employees']}")
            if org.get("industry"):
                notas_extra.append(f"Sector: {org['industry']}")
            if org.get("short_description"):
                notas_extra.append(f"Descripción: {org['short_description'][:200]}...")
            
            if notas_extra:
                current_notas = lead.get("notas") or ""
                if "Apollo Data" not in current_notas:
                    update_data["notas"] = f"{current_notas}\n\n--- Apollo Data ---\n" + "\n".join(notas_extra)
            
            if update_data:
                await db.leads.update_one({"lead_id": lead_id}, {"$set": update_data})
            
            return {
                "success": True,
                "updated_fields": list(update_data.keys()),
                "telefono": org.get("phone"),
                "cargo": None,
                "sector": org.get("industry"),
                "notas": update_data.get("notas", "")
            }
        else:
            raise HTTPException(status_code=response.status_code, detail=f"Apollo API error: {response.text}")


@api_router.get("/apollo/health")
async def apollo_health_check(current_user: UserResponse = Depends(get_current_user)):
    if not APOLLO_API_KEY:
        return {"status": "error", "message": "Apollo API key not configured"}
    headers = {"Content-Type": "application/json", "X-Api-Key": APOLLO_API_KEY}
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(f"{APOLLO_BASE_URL}/organizations/enrich", headers=headers, params={"domain": "apollo.io"}, timeout=10.0)
            return {"status": "connected" if response.status_code == 200 else "error", "message": "Apollo API working" if response.status_code == 200 else f"Apollo returned {response.status_code}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@api_router.get("/")
async def root():
    return {"message": "System Rapid Solutions CRM API", "version": "1.1.0"}

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
