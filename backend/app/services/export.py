"""Export de oportunidades a Excel."""
from typing import List, Dict, Optional
from datetime import datetime
import io

def export_to_excel(oportunidades: List[Dict], filename: str = None) -> bytes:
    """Exportar oportunidades a Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Oportunidades"
    
    # Headers
    headers = [
        "Expediente", "Tipo", "Sector", "Score", "Nivel",
        "Título", "Órgano Contratación", "Importe",
        "Días Restantes", "Estado", "Propietario",
        "CPV", "Categoría", "Adjudicatario", "URL"
    ]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    # Escribir datos
    for row, op in enumerate(oportunidades, 2):
        ws.cell(row=row, column=1, value=op.get("expediente"))
        ws.cell(row=row, column=2, value=op.get("tipo"))
        ws.cell(row=row, column=3, value=op.get("sector"))
        ws.cell(row=row, column=4, value=op.get("scoring", {}).get("total", 0))
        ws.cell(row=row, column=5, value=op.get("scoring", {}).get("nivel", ""))
        ws.cell(row=row, column=6, value=op.get("titulo", "")[:100])
        ws.cell(row=row, column=7, value=op.get("organo_contratacion", ""))
        ws.cell(row=row, column=8, value=op.get("importe"))
        ws.cell(row=row, column=9, value=op.get("dias_restantes"))
        ws.cell(row=row, column=10, value=op.get("estado"))
        ws.cell(row=row, column=11, value=op.get("propietario_nombre", ""))
        ws.cell(row=row, column=12, value=op.get("cpv", ""))
        ws.cell(row=row, column=13, value=op.get("analisis", {}).get("categoria", ""))
        ws.cell(row=row, column=14, value=op.get("adjudicatario", {}).get("nombre", "") if op.get("adjudicatario") else "")
        ws.cell(row=row, column=15, value=op.get("url_licitacion", ""))
        
        # Color por nivel
        nivel = op.get("scoring", {}).get("nivel", "")
        if nivel == "oro":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="FFF2CC", fill_type="solid")
        elif nivel == "plata":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", fill_type="solid")
    
    # Ajustar anchos
    column_widths = [15, 12, 10, 8, 8, 50, 30, 15, 12, 15, 20, 12, 15, 30, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def export_to_csv(oportunidades: List[Dict]) -> str:
    """Exportar oportunidades a CSV."""
    import csv
    import io
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers
    headers = [
        "expediente", "tipo", "sector", "score", "nivel",
        "titulo", "organo_contratacion", "importe",
        "dias_restantes", "estado", "propietario",
        "cpv", "categoria", "adjudicatario", "url"
    ]
    writer.writerow(headers)
    
    # Datos
    for op in oportunidades:
        writer.writerow([
            op.get("expediente"),
            op.get("tipo"),
            op.get("sector"),
            op.get("scoring", {}).get("total", 0),
            op.get("scoring", {}).get("nivel", ""),
            op.get("titulo", ""),
            op.get("organo_contratacion", ""),
            op.get("importe"),
            op.get("dias_restantes"),
            op.get("estado"),
            op.get("propietario_nombre", ""),
            op.get("cpv", ""),
            op.get("analisis", {}).get("categoria", ""),
            op.get("adjudicatario", {}).get("nombre", "") if op.get("adjudicatario") else "",
            op.get("url_licitacion", "")
        ])
    
    return output.getvalue()
